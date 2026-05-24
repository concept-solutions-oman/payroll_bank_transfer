# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, api, _

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

class BankTransferWizard(models.TransientModel):
    _name = 'bank.transfer.wizard'
    _description = 'Bank Transfer Excel Export'

    payslip_ids = fields.Many2many(
        'hr.payslip',
        string='Payslips',
    )
    employer_cr_no = fields.Char(string='Employer CR-NO', default='5086301', readonly=True)
    payer_cr_no = fields.Char(string='Payer CR-NO', default='5086301', readonly=True)
    payer_bank_short_name = fields.Char(string='Payer Bank', default='OAB', readonly=True)
    payer_account_number = fields.Char(string='Payer A/c Number', default='3160684788502', readonly=True)
    payment_type = fields.Selection([
        ('Monthly Salary', 'Monthly Salary'),
        ('Late Salary Payment', 'Late Salary Payment'),
        ('Advance Salary Payment', 'Advance Salary Payment'),
        ('Bonus', 'Bonus'),
        ('Overtime Payment', 'Overtime Payment'),
        ('Allowance', 'Allowance'),
        ('EndofserviceBenefit', 'EndofserviceBenefit'),
    ], string='Payment Type', default='Monthly Salary', required=True)
    
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_ids = self.env.context.get('active_ids', [])
        if active_ids:
            res['payslip_ids'] = [(6, 0, active_ids)]
        return res

    def action_export_excel(self):
        self.ensure_one()
        if not Workbook:
            raise models.ValidationError("The 'openpyxl' Python library is required.")
        
        payslips = self.payslip_ids
        if not payslips:
            raise models.ValidationError("No payslips selected for export.")
            
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bank Transfer'
        
        # Styles
        bold_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        
        # Calculate summary values
        total_salaries = 0.0
        
        salary_month = ''
        salary_year = ''
        if payslips:
            first_date = payslips[0].date_to
            if first_date:
                salary_month = first_date.strftime('%m')
                salary_year = first_date.strftime('%Y')
                
        employee_data = []
        for payslip in payslips:
            emp = payslip.employee_id
            
            # Extract BIC & Account
            emp_account = ''
            emp_bic = ''
            if emp.bank_account_id:
                import re
                emp_account = re.sub(r'[^a-zA-Z0-9]', '', emp.bank_account_id.acc_number or '')
                if emp.bank_account_id.bank_id:
                    emp_bic = emp.bank_account_id.bank_id.bic or emp.bank_account_id.bank_id.name or ''
                    emp_bic = emp_bic.upper()
            
            # Calculate Salary, Income, Deductions, SPF
            net_salary = 0.0
            basic_salary = 0.0
            total_allowance = 0.0
            total_deduction = 0.0
            spf_deduction = 0.0
            working_days = 0.0
            
            for line in payslip.line_ids:
                cat_code = line.category_id.code if line.category_id else ''
                # Skip structural/subtotal lines to avoid double-counting
                # SUBTOTAL lines (TOT_ALW, TOT_OT, TOT_DED) are sums of individual lines already counted
                if cat_code in ('BASIC', 'GROSS', 'NET', 'SUBTOTAL'):
                    if cat_code == 'BASIC':
                        basic_salary += line.total
                        if line.quantity > 0:
                            working_days = line.quantity
                    elif cat_code == 'NET':
                        net_salary += line.total
                    continue
                # Deductions
                if cat_code == 'DED':
                    line_code_lower_ded = (line.code or '').lower()
                    line_name_lower_ded = (line.name or '').lower()
                    if 'spf' in line_code_lower_ded or 'spf' in line_name_lower_ded or 'social security' in line_name_lower_ded:
                        spf_deduction += abs(line.total)
                    else:
                        total_deduction += abs(line.total)
                # Extra income: ANY positive category that isn't BASIC/GROSS/NET/SUBTOTAL/DED/COMP
                # This includes ALW, OT, DDA (Desert Daily Rate), and any other custom allowance categories
                else:
                    if cat_code not in ('COMP', '') and line.total > 0:
                        # Exclude 'Attendance Pay' or 'No Attendance' lines from extra income
                        line_name_lower = (line.name or '').lower()
                        line_code_lower = (line.code or '').lower()
                        if 'attendance' in line_name_lower or 'attendance' in line_code_lower:
                            continue
                        
                        total_allowance += line.total
            
            # Extra Hours total and Working Days (from Worked Days)
            extra_hours = 0.0
            for wd in payslip.worked_days_line_ids:
                if wd.code == 'OT': # Usually Overtime is OT
                    extra_hours += wd.number_of_hours
            
            working_days = int(working_days)
            if working_days == 0:
                continue
                    
            total_salaries += round(net_salary, 3)
            
            # Month/Year for notes
            month_year_note = ''
            ref_no = ''
            if payslip.date_to:
                month_year_note = payslip.date_to.strftime('%b %Y Salary') # Example: Jan 2026 Salary
                ref_no = emp.employee_number or ''
                
            import re
            clean_name = re.sub(r'[^a-zA-Z0-9\s]', '', emp.name or '')
            clean_name = clean_name[:30]  # Max 30 characters as per bank spec
            
            nationality_val = 'EXPAT'
            if emp.country_id and emp.country_id.name and emp.country_id.name.upper() == 'OMAN':
                nationality_val = 'OMAN'

            # Fetch division from department
            division_name = emp.department_id.name or ''
                
            employee_data.append({
                'id_type': 'C',
                'id_no': emp.identification_id or '',
                'ref_no': ref_no,
                'name': clean_name,
                'nationality': nationality_val,
                'division': division_name,
                'bic': emp_bic,
                'account': emp_account,
                'freq': 'M',
                'working_days': working_days,
                'net_salary': round(net_salary, 3),
                'basic_salary': round(basic_salary, 3),
                'extra_hours': 0,
                'extra_income': round(total_allowance, 3),
                'deductions': round(total_deduction, 3),
                'social_security': round(spf_deduction, 3),
                'notes': month_year_note,
            })
            
        number_of_records = len(employee_data)
        
        # Headers
        headers = [
            'Employee ID Type', 'Employee ID', 'Reference Number', 'Employee Name',
            'Employee BIC', 'Employee Account', 'Salary Frequency', 'Working days', 'Net Salary',
            'Basic Salary', 'Extra hours', 'Extra income', 'Deductions', 'Social Security Deductions', 'Notes / Comments', 'Nationality', 'Division'
        ]
        ws.append(headers)
        center_alignment = Alignment(horizontal='center', vertical='center')
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            
        # Row 4+ Data
        for data in employee_data:
            row_data = [
                data['id_type'],
                data['id_no'],
                data['ref_no'],
                data['name'],
                data['bic'],
                data['account'],
                data['freq'],
                data['working_days'], # Working days logic
                data['net_salary'],
                data['basic_salary'],
                data['extra_hours'],
                data['extra_income'],
                data['deductions'],
                data['social_security'],
                data['notes'],
                data['nationality'],
                data['division']
            ]
            ws.append(row_data)
            # Apply 3-decimal number format to all monetary columns
            current_row = ws.max_row
            money_cols = [9, 10, 12, 13, 14]  # Net, Basic, Extra Income, Deductions, Social Security
            for col_idx in money_cols:
                ws.cell(row=current_row, column=col_idx).number_format = '0.000'
            
        # Set column widths
        for col in range(1, 18):
            ws.column_dimensions[get_column_letter(col)].width = 20
            
        # Output Excel
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.excel_file = base64.b64encode(output.read())
        self.file_name = 'Bank_Transfer_Report.xlsx'
        
        return {
            'type': 'ir.actions.act_window',
            'name': 'Download Bank Transfer Report',
            'res_model': 'bank.transfer.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': self.env.context,
        }
